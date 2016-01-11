import os
import glob
glob.__file__
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

# path
spath = "C:\\Users\\Christopher.Tully\\Documents\\Excel Wrap Scripts\\Test"

# create wb in memory to write files to
dest_wb = Workbook()
ws = dest_wb.active
ws.title = "QA Scores"

# create list & iterator to hold name & scores
name_scores = []
x = 0

# open directory and read specified files
os.chdir(spath)
for file in glob.glob("*/2016/December/*.xlsx"):
    
    # Loop through excel files & copy cells A1 & B1
    wb = load_workbook(file, data_only=True)
    front = wb.get_sheet_by_name('Sheet1')

    # copy scores to 'name_score' list
    for i in range (1, 3):
        name_scores.append(front.cell(row=1, column=i).value)

# iterate over length of list and paste them to new wb
while (x < len(name_scores)):
    for rowNum in range(1,3):
        for columnNum in range (1,3):
            ws_cell = ws.cell(row = rowNum, column=columnNum)
            ws_cell.value = name_scores[x]
            x += 1

# save to final wb
final_wb = "C:\\Users\\Christopher.Tully\\Documents\\Excel Wrap Scripts\\Test"
os.chdir(final_wb)
dest_wb.save("Wrap Up.xlsx")

